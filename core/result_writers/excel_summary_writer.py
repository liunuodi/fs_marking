# core/result_writers/excel_summary_writer.py
import os
from openpyxl import load_workbook

class ExcelSummaryWriter:
    def __init__(
        self,
        template_path: str,
        output_path:   str,
        sheet_name:    str,
        zid_row:       int,
        rule_rows:     dict,   # {"margin":4, "cover_page":5, ...}
        total_row:     int,
    ):
        self.template_path = template_path
        self.output_path   = output_path
        self.sheet_name    = sheet_name
        self.zid_row       = zid_row
        self.rule_rows     = rule_rows
        self.total_row     = total_row
        self.cache         = {}    # {zid: result_dict}

    # 收集单个学生结果
    def write(self, zid: str, result: dict):
        self.cache[str(zid)] = result

    def save(self):
        folder = os.path.dirname(self.output_path)
        if folder:
            os.makedirs(folder, exist_ok=True)

        wb = load_workbook(self.template_path)
        ws = wb[self.sheet_name]

        # 1) ZID -> col
        zid2col = {
            str(cell.value): cell.col_idx
            for cell in ws[self.zid_row] 
            if cell.value
        }

        # 2) 写入
        for zid, res in self.cache.items():
            col = zid2col.get(zid)
            if not col:
                print(f"⚠️ ZID {zid} 不在模板中，跳过")
                continue

            # 写规则分
            for item in res["results"]:
                rule = item["name"].lower()
                mark = item["mark"]
                row  = self.rule_rows.get(rule)
                if row:
                    ws.cell(row=row, column=col, value=mark)
                else:
                    print(f"⚠️ 模板缺少规则行 {rule}")

            # 写总分
            if self.total_row:
                ws.cell(row=self.total_row, column=col, value=res["total"])

        wb.save(self.output_path)
        print(f"✔️ 结果已写入 {self.output_path}")